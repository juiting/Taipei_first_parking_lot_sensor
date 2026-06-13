"""Excel 匯出：車格明細（狀態/停入離場時間/電量/RSSI）+ 統計摘要。

色階與門檻與前端一致（電量六段、RSSI 四級），維護時兩邊要同步調整。
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STATUS_ZH = {
    "Available": "空位", "Occupied": "在席", "Allocated": "已分配",
    "Maintenance": "維護", "Unknown": "無資料",
}

# Excel 淡色系填色（ARGB）
_FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
_FILL_LIGHTGREEN = PatternFill("solid", fgColor="E2F0D9")
_FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_FILL_LIGHTRED = PatternFill("solid", fgColor="FFD8C9")
_FILL_RED = PatternFill("solid", fgColor="FFC7CE")
_FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
_FILL_HEADER = PatternFill("solid", fgColor="1F2937")


def _battery_label_fill(b: int | None) -> tuple[str, PatternFill | None]:
    if b is None:
        return "—", None
    if b >= 95:
        return f"{b}%", _FILL_GREEN
    if b >= 80:
        return f"{b}%", _FILL_LIGHTGREEN
    if b >= 50:
        return f"{b}%", _FILL_YELLOW
    if b >= 30:
        return f"{b}%", _FILL_LIGHTRED
    return f"{b}%", _FILL_RED


def _rssi_grade(r: int | None) -> str:
    if r is None:
        return "無資料"
    if r >= -85:
        return "優"
    if r >= -95:
        return "良"
    if r >= -105:
        return "中"
    return "差"


_RSSI_FILL = {"優": _FILL_GREEN, "良": _FILL_LIGHTGREEN, "中": _FILL_YELLOW, "差": _FILL_RED, "無資料": _FILL_GRAY}
_STATUS_FILL = {"空位": _FILL_GREEN, "在席": _FILL_RED, "維護": _FILL_GRAY, "已分配": _FILL_YELLOW, "無資料": _FILL_GRAY}

_HEADERS = [
    "車格", "類型", "分區", "標籤", "狀態", "感測器離線",
    "停入時間", "離場時間", "最後回報",
    "電量(%)", "電量更新時間", "RSSI(dBm)", "訊號等級", "RSSI量測時間",
    "Carin IMEI", "Brickcom MAC",
]
_COL_WIDTHS = [9, 8, 7, 10, 9, 10, 19, 19, 19, 9, 19, 10, 9, 19, 13, 19]


def _sort_key(space: dict):
    name = space["name"]
    return (0, int(name)) if name.isdigit() else (1, name)


def build_workbook(spaces: list[dict], health: dict[str, dict], site_name: str) -> bytes:
    now = datetime.now().replace(microsecond=0)
    wb = Workbook()

    # ---- 工作表 1：車格明細 ----
    ws = wb.active
    ws.title = "車格明細"
    ws.append(_HEADERS)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTHS[i - 1]
    ws.freeze_panes = "A2"

    counts = {"空位": 0, "在席": 0, "維護": 0, "已分配": 0, "無資料": 0}
    offline = 0
    bat_buckets = {"≥95%": 0, "80–95%": 0, "50–80%": 0, "30–50%": 0, "10–30%": 0, "<10%": 0, "無資料": 0}
    rssi_buckets = {"優": 0, "良": 0, "中": 0, "差": 0, "無資料": 0}

    for s in sorted(spaces, key=_sort_key):
        status_zh = STATUS_ZH.get(s.get("status", "Unknown"), "無資料")
        counts[status_zh] = counts.get(status_zh, 0) + 1
        is_offline = bool(s.get("offline"))
        offline += is_offline

        occupied = s.get("status") == "Occupied"
        available = s.get("status") == "Available"
        event_time = s.get("event_time") or ""

        battery = s.get("battery")
        if battery is None:
            bat_buckets["無資料"] += 1
        elif battery >= 95:
            bat_buckets["≥95%"] += 1
        elif battery >= 80:
            bat_buckets["80–95%"] += 1
        elif battery >= 50:
            bat_buckets["50–80%"] += 1
        elif battery >= 30:
            bat_buckets["30–50%"] += 1
        elif battery >= 10:
            bat_buckets["10–30%"] += 1
        else:
            bat_buckets["<10%"] += 1

        rssi = s.get("rssi")
        grade = _rssi_grade(rssi)
        rssi_buckets[grade] += 1

        h = health.get(s["name"], {})
        ws.append([
            s["name"],
            "大客車" if s.get("type") == "bus" else "小型車",
            s.get("zone") or "",
            " ".join(s.get("tags") or []),
            status_zh,
            "離線" if is_offline else "",
            event_time if occupied else "",
            event_time if available else "",
            s.get("last_heartbeat") or "",
            battery if battery is not None else "—",
            s.get("battery_at") or "",
            rssi if rssi is not None else "—",
            grade,
            s.get("rssi_at") or "",
            s.get("imei") or "",
            h.get("mac") or "",
        ])
        row = ws.max_row
        ws.cell(row=row, column=5).fill = _STATUS_FILL.get(status_zh, _FILL_GRAY)
        if is_offline:
            ws.cell(row=row, column=6).font = Font(color="9C0006", bold=True)
        _, bfill = _battery_label_fill(battery)
        if bfill:
            ws.cell(row=row, column=10).fill = bfill
            if battery is not None and battery < 10:
                ws.cell(row=row, column=10).font = Font(color="9C0006", bold=True)
        ws.cell(row=row, column=13).fill = _RSSI_FILL[grade]

    # ---- 工作表 2：統計摘要 ----
    ws2 = wb.create_sheet("統計摘要")
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 12

    def section(title: str, items: dict):
        ws2.append([title])
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, size=12)
        for k, v in items.items():
            ws2.append([k, v])
        ws2.append([])

    ws2.append([f"{site_name} 車格狀態匯出"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.append(["產出時間", str(now)])
    ws2.append(["總格數", len(spaces)])
    ws2.append([])
    section("停車狀態", {**{k: v for k, v in counts.items() if v}, "感測器離線": offline})
    section("電量分布", bat_buckets)
    section("訊號分布（RSSI）", rssi_buckets)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_ISSUE_FILL = {
    "mac_mismatch": PatternFill("solid", fgColor="FFC7CE"),
    "offline": PatternFill("solid", fgColor="D9D9D9"),
    "low_battery": PatternFill("solid", fgColor="FFD8C9"),
    "weak_signal": PatternFill("solid", fgColor="FFEB9C"),
    "remark_flag": PatternFill("solid", fgColor="E2EFDA"),
}
_ISSUE_HEADERS = [
    "車格", "分區", "主要問題", "問題類別", "Carin MAC", "Brickcom MAC",
    "電量(%)", "RSSI(dBm)", "Brickcom 備註", "說明",
]
_ISSUE_WIDTHS = [9, 7, 16, 24, 14, 18, 9, 10, 26, 40]


def build_issues_workbook(report: dict, site_name: str) -> bytes:
    from .issues import ISSUE_META

    wb = Workbook()

    # ---- 摘要 ----
    ws0 = wb.active
    ws0.title = "異常摘要"
    ws0.column_dimensions["A"].width = 18
    ws0.column_dimensions["B"].width = 10
    ws0.column_dimensions["C"].width = 44
    ws0.append([f"{site_name} 設備異常報表"])
    ws0.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws0.append(["產出時間", report["generated_at"]])
    ws0.append(["異常車格總數", report["total_issues"]])
    if not report["brickcom_enabled"]:
        ws0.append(["備註", "未啟用 Brickcom，僅統計 Carin 可判斷之斷線類"])
    ws0.append([])
    ws0.append(["問題類別", "車格數", "處理建議"])
    for c in ws0[ws0.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = _FILL_HEADER
    for cat in report["categories"]:
        ws0.append([cat["label"], cat["count"], cat["hint"]])
        ws0.cell(row=ws0.max_row, column=1).fill = _ISSUE_FILL.get(cat["key"], _FILL_GRAY)

    # ---- 明細（依嚴重度排序）----
    ws = wb.create_sheet("異常明細")
    ws.append(_ISSUE_HEADERS)
    for i, cell in enumerate(ws[1], start=1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = _ISSUE_WIDTHS[i - 1]
    ws.freeze_panes = "A2"

    for it in report["items"]:
        labels = "、".join(ISSUE_META[f]["label"] for f in it["flags"])
        detail = "；".join(it["detail"].get(f, "") for f in it["flags"] if it["detail"].get(f))
        ws.append([
            it["name"], it.get("zone") or "",
            ISSUE_META[it["primary"]]["label"], labels,
            it.get("imei") or "", it.get("mac") or "",
            it["battery"] if it["battery"] is not None else "—",
            it["rssi"] if it["rssi"] is not None else "—",
            it.get("remark") or "", detail,
        ])
        ws.cell(row=ws.max_row, column=3).fill = _ISSUE_FILL.get(it["primary"], _FILL_GRAY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
